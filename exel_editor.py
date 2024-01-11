import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


#open
workbook = openpyxl.load_workbook("draft_1.xlsx")
data_row = workbook["Raw data (last 100 scans)"]
data_without_noise = workbook["Noise removal"]
reference_row =  workbook["Reference (last 100 scans)"]
reference_without_noise = workbook["Ref noise removal"]
comparison = workbook["Comparison"]
plots = workbook["Barchart"]
result_sheet = workbook["Result sheet"]


def raw_input(data_samples,type): 


        # Записываем данные в Excel
    for row_idx, row_dict in enumerate(data_samples, start=2):
        probe_number = row_dict.get('Номер пробы', "?")
        existing_row = None
        for existing_row_idx, existing_row_dict in enumerate(data_samples, start=2):
            if existing_row_dict.get('Номер пробы') == probe_number:
                existing_row = existing_row_idx
                break

        if existing_row is None:
            existing_row = len(data_row['D']) + 1 
        
        # number of probes 
        if  type == "data": 
            data_row.cell(row=existing_row, column=4, value=probe_number)
            data_without_noise.cell(row=existing_row, column=4, value=probe_number)
        if  type == "reference":
            reference_row.cell(row=existing_row, column=4, value=probe_number)
            reference_without_noise.cell(row=existing_row, column=4, value=probe_number)

        # write a meassurment in cells 
        for col_idx, value in enumerate(row_dict.get('Значения пробы', []), start=5):  # start with E (5)
            if  type == "data":
                data_row.cell(row=existing_row, column=col_idx, value=value)
                data_without_noise.cell(row=existing_row, column=col_idx, value=value if (value > 0) else 0)
                max_row = existing_row
                max_col = col_idx
            if  type == "reference":
                reference_row.cell(row=existing_row, column=col_idx, value=value)
                reference_without_noise.cell(row=existing_row, column=col_idx, value=value if (value > 0) else 0)
                max_row = existing_row
                max_col = col_idx
        print (type, max_row,max_col)
    
    for col_idx in range(5, max_col + 1):
        average = f"=AVERAGE({get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{max_row})"
        if type == "data":
            data_row.cell(row=max_row +1, column=col_idx, value=average)
            data_without_noise.cell(row=max_row + 1, column=col_idx, value=average) 
        if type == "reference":
            reference_row.cell(row=max_row +1, column=col_idx, value=average)
            reference_without_noise.cell(row=max_row + 1, column=col_idx, value=average) 
    
    #sum of mass 

    for row_idx in range (2, max_row + 2): 
        sum = f"=SUM(E{row_idx}:{get_column_letter(max_col)}{row_idx})"
        if type == "data":
            data_row.cell(row=row_idx, column=1, value=sum)
            data_without_noise.cell(row=row_idx, column=1, value=sum) 
        if type == "reference":
            reference_row.cell(row=row_idx, column=1, value=sum)
            reference_without_noise.cell(row=row_idx, column=1, value=sum) 

    # sum of mass > 44: 
    value_to_find = 45
    column_letter = None

    for col_idx in range(1, data_row.max_column + 1):
        if type == "data":
            cell_value = data_row.cell(row=1, column=col_idx).value
            cell_value_without_noise = data_without_noise.cell(row=1, column=col_idx).value
        if type == "reference":
            cell_value = reference_row.cell(row=1, column=col_idx).value
            cell_value_without_noise = reference_without_noise.cell(row=1, column=col_idx).value
        if cell_value == value_to_find:
            mass45 = get_column_letter(col_idx)
            break
        if cell_value_without_noise == value_to_find:
            mass45 = get_column_letter(col_idx)
            break

    for row_idx in range (2, max_row + 2): 
        sum = f"=SUM({mass45}{row_idx}:{get_column_letter(max_col)}{row_idx})"
        if type == "data":        
            data_row.cell(row=row_idx, column=2, value=sum)
            data_without_noise.cell(row=row_idx, column=2, value=sum)
        if type == "reference":        
            reference_row.cell(row=row_idx, column=2, value=sum)
            reference_without_noise.cell(row=row_idx, column=2, value=sum)

    for row_idx in range (2, max_row + 2): 
        ratio = f"= B{row_idx}/A{row_idx}"
        if type == "data":            
            data_row.cell(row=row_idx, column = 3,value = ratio)
            data_without_noise.cell(row=row_idx, column = 3,value = ratio)
        if type == "reference":            
            reference_row.cell(row=row_idx, column = 3,value = ratio)
            reference_without_noise.cell(row=row_idx, column = 3,value = ratio)

    value_to_find = "18 (H20)"
    column_letter = None

    mass18 = None  # Инициализируем mass18 перед использованием

    for col_idx in range(1, data_row.max_column + 1):
        if type == "data":            
           cell_value = data_row.cell(row=1, column=col_idx).value
        if type == "reference":            
           cell_value = reference_row.cell(row=1, column=col_idx).value    
        if cell_value == value_to_find:
            mass18 = get_column_letter(col_idx)
            break

    for col_idx in range(5, max_col + 1):
        check18_100= f"=IF({get_column_letter(col_idx)}{max_row+1} < {mass18}{max_row+1}/100, 0, 1)"
        if type == "data":     
            data_row.cell(row=max_row + 2, column=col_idx, value=check18_100)
        check18_1000= f"=IF({get_column_letter(col_idx)}{max_row+1} < {mass18}{max_row+1}/1000, 0, 1)"
        if type == "data":    
            data_row.cell(row=max_row + 3, column=col_idx, value=check18_1000)
#comparison list 
    #value (data - reference)
    if type == "reference":       
        for row_idx in range(2, max_row + 1):
            for col_idx in range(5, max_col + 1):
                data_value = data_without_noise.cell(row=row_idx, column=col_idx).value
                reference_value = reference_without_noise.cell(row=row_idx, column=col_idx).value
                difference = data_value - reference_value
                comparison.cell(row=row_idx, column=col_idx, value=difference)
        for col_idx in range(5, max_col + 1):
            average = f"=AVERAGE({get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{max_row})"
            comparison.cell(row=max_row +1, column=col_idx, value=average)
    #summ of all 
        for row_idx in range (2, max_row + 2): 
            sum = f"=SUM(E{row_idx}:{get_column_letter(max_col)}{row_idx})"
            comparison.cell(row=row_idx, column=1, value=sum) 
        for col_idx in range(1, data_row.max_column + 1):
            cell_value = comparison.cell(row=1, column=col_idx).value
            value_to_find = 45
            if cell_value == value_to_find:
                mass45 = get_column_letter(col_idx)
                break
        for row_idx in range (2, max_row + 2): 
            sum = f"=SUM({mass45}{row_idx}:{get_column_letter(max_col)}{row_idx})"      
            comparison.cell(row=row_idx, column=2, value=sum)
        for row_idx in range (2, max_row + 2): 
            ratio = f"= B{row_idx}/A{row_idx}"       
            comparison.cell(row=row_idx, column = 3,value = ratio)
   
    max_row = 0 
    max_col = 0

    if type == "reference":
        workbook.save('test1.xlsx')